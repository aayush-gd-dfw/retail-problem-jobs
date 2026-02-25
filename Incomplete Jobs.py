import io
import re
import base64
import os
from datetime import datetime
from typing import Optional, Dict, Any, List, Tuple

import requests
from msal import ConfidentialClientApplication
from openpyxl import load_workbook

GRAPH = "https://graph.microsoft.com/v1.0"

# =========================================================
# HARD-CODED AUTH CONFIG (EDIT THESE 3)
# =========================================================

TENANT_ID = os.getenv("TENANT_ID")
CLIENT_ID = os.getenv("CLIENT_ID")
CLIENT_SECRET = os.getenv("CLIENT_SECRET")
# =========================================================

# -------------------- CONFIG --------------------
SUBJECT = "Retail Problem/Incomplete Jobs Report"

# SharePoint list column internal names
# IMPORTANT: AssignedTo is now a TEXT column (single line of text)
FIELDS = {
    "job_number": "Title",            # display "Job Number" -> internal "Title"
    "customer_name": "CustomerName",
    "job_amount": "JobAmount",
    "next_appt_date": "NextApptDate",
    "assigned_to": "EmployeeName",      # TEXT column now
}

# BU -> assigned email mapping (normalize_key() is used)
ASSIGNED_MAP = {
    "dallas": "amunoz@GlassDoctorDFW.com",
    "carrollton": "amunoz@GlassDoctorDFW.com",
    "arlington": "ayeamans@glassdoctordfw.com",
    "colleyville": "ayeamans@glassdoctordfw.com",
    "denton": "ayeamans@glassdoctordfw.com"
}

# -------------------- Small helpers --------------------
def parse_dt(dt_str: str) -> datetime:
    if not dt_str:
        return datetime(1970, 1, 1)
    if dt_str.endswith("Z"):
        dt_str = dt_str.replace("Z", "+00:00")
    return datetime.fromisoformat(dt_str)

def normalize_key(s: str) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip().lower())

def find_col_idx(header: List[Any], wanted: List[str]) -> Optional[int]:
    h = [normalize_key(x) for x in header]
    wanted_norm = [normalize_key(w) for w in wanted]
    for i, name in enumerate(h):
        if name in wanted_norm:
            return i
    for i, name in enumerate(h):
        for w in wanted_norm:
            if w in name:
                return i
    return None

def parse_money(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"[^0-9.\-]", "", str(v))
    return float(s) if s else 0.0

def try_parse_excel_date(v) -> Optional[str]:
    """
    SharePoint Date/Time via Graph prefers ISO datetime strings:
    YYYY-MM-DDT00:00:00Z
    """
    if v is None or str(v).strip() == "":
        return None

    if isinstance(v, datetime):
        d = v.date().isoformat()
        return f"{d}T00:00:00Z"

    try:
        s = str(v).strip()
        for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%Y/%m/%d"):
            try:
                d = datetime.strptime(s, fmt).date().isoformat()
                return f"{d}T00:00:00Z"
            except Exception:
                pass

        d = datetime.fromisoformat(s).date().isoformat()
        return f"{d}T00:00:00Z"
    except Exception:
        return None

# -------------------- Auth --------------------
def get_graph_token() -> str:
    app = ConfidentialClientApplication(
        client_id=CLIENT_ID,
        client_credential=CLIENT_SECRET,
        authority=f"https://login.microsoftonline.com/{TENANT_ID}",
    )
    result = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    if "access_token" not in result:
        raise RuntimeError(f"Token error: {result.get('error')} - {result.get('error_description')}")
    return result["access_token"]

# -------------------- Graph HTTP helpers --------------------
def graph_get(token: str, url: str, params: Optional[dict] = None) -> Dict[str, Any]:
    headers = {"Authorization": f"Bearer {token}", "ConsistencyLevel": "eventual"}
    r = requests.get(url, headers=headers, params=params, timeout=60)
    if not r.ok:
        try:
            print("Graph error:", r.json())
        except Exception:
            print("Graph error:", r.text)
        r.raise_for_status()
    return r.json()

def graph_get_bytes(token: str, url: str) -> bytes:
    headers = {"Authorization": f"Bearer {token}", "ConsistencyLevel": "eventual"}
    r = requests.get(url, headers=headers, timeout=120)
    if not r.ok:
        try:
            print("Graph error:", r.json())
        except Exception:
            print("Graph error:", r.text)
        r.raise_for_status()
    return r.content

def graph_post(token: str, url: str, payload: dict) -> Dict[str, Any]:
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    r = requests.post(url, headers=headers, json=payload, timeout=60)
    if not r.ok:
        print("POST url:", url)
        print("POST payload:", payload)
        try:
            print("Graph error JSON:", r.json())
        except Exception:
            print("Graph error TEXT:", r.text)
        r.raise_for_status()
    return r.json()

def graph_patch(token: str, url: str, payload: dict) -> None:
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    r = requests.patch(url, headers=headers, json=payload, timeout=60)
    if not r.ok:
        print("PATCH url:", url)
        print("PATCH payload:", payload)
        try:
            print("Graph error JSON:", r.json())
        except Exception:
            print("Graph error TEXT:", r.text)
        r.raise_for_status()

# -------------------- Outlook: latest email + attachment --------------------
def latest_message_for_subject(token: str, mailbox_upn: str, subject_phrase: str) -> Optional[Dict[str, Any]]:
    url = f"{GRAPH}/users/{mailbox_upn}/mailFolders/Inbox/messages"
    params = {
        "$select": "id,subject,receivedDateTime,hasAttachments",
        "$top": "25",
        "$search": f"\"{subject_phrase}\"",
    }
    data = graph_get(token, url, params=params)
    msgs: List[Dict[str, Any]] = data.get("value", [])
    phrase = subject_phrase.lower()
    candidates = [m for m in msgs if phrase in (m.get("subject") or "").lower()]
    if not candidates:
        return None
    candidates.sort(key=lambda m: parse_dt(m.get("receivedDateTime", "")), reverse=True)
    return candidates[0]

def get_first_xlsx_attachment_from_message(token: str, mailbox_upn: str, message_id: str) -> Tuple[Optional[str], Optional[bytes]]:
    url = f"{GRAPH}/users/{mailbox_upn}/messages/{message_id}/attachments"
    data = graph_get(token, url, params={"$top": "50"})
    for a in data.get("value", []):
        name = a.get("name") or ""
        if name.lower().endswith(".xlsx"):
            cb = a.get("contentBytes")
            if cb:
                return name, base64.b64decode(cb)
            att_id = a.get("id")
            raw_url = f"{GRAPH}/users/{mailbox_upn}/messages/{message_id}/attachments/{att_id}/$value"
            return name, graph_get_bytes(token, raw_url)
    return None, None

# -------------------- XLSX parsing --------------------
def read_xlsx_first_sheet_rows(xlsx_bytes: bytes) -> List[List[Any]]:
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for r in ws.iter_rows(values_only=True):
        rows.append([("" if v is None else v) for v in r])
    return rows

# -------------------- SharePoint list discovery + upsert --------------------
def get_site_and_list_ids(token: str, hostname: str, site_path: str, list_display_name: str) -> Tuple[str, str]:
    site = graph_get(token, f"{GRAPH}/sites/{hostname}:{site_path}")
    site_id = site["id"]

    lists = graph_get(token, f"{GRAPH}/sites/{site_id}/lists", params={"$select": "id,displayName"})
    for l in lists.get("value", []):
        if (l.get("displayName") or "").strip().lower() == list_display_name.strip().lower():
            return site_id, l["id"]
    raise RuntimeError(f"List not found: {list_display_name}")

def load_existing_job_map(token: str, site_id: str, list_id: str, job_field_internal: str) -> Dict[str, str]:
    """
    No $filter. Build {job_number: item_id}.
    """
    url = f"{GRAPH}/sites/{site_id}/lists/{list_id}/items"
    params = {"$expand": f"fields($select={job_field_internal})", "$top": "200"}

    job_to_id: Dict[str, str] = {}

    while True:
        data = graph_get(token, url, params=params)
        for it in data.get("value", []):
            item_id = it.get("id")
            fields = it.get("fields", {}) or {}
            job_val = fields.get(job_field_internal)

            if not item_id or job_val in (None, ""):
                continue

            key = str(job_val).strip()
            if key:
                job_to_id[key] = item_id

        next_link = data.get("@odata.nextLink")
        if not next_link:
            break
        url = next_link
        params = None

    return job_to_id

def create_item(token: str, site_id: str, list_id: str, fields_payload: dict) -> Dict[str, Any]:
    url = f"{GRAPH}/sites/{site_id}/lists/{list_id}/items"
    return graph_post(token, url, {"fields": fields_payload})

def update_item_fields(token: str, site_id: str, list_id: str, item_id: str, fields_payload: dict) -> None:
    url = f"{GRAPH}/sites/{site_id}/lists/{list_id}/items/{item_id}/fields"
    graph_patch(token, url, fields_payload)

def print_list_columns(token: str, site_id: str, list_id: str) -> None:
    cols = graph_get(token, f"{GRAPH}/sites/{site_id}/lists/{list_id}/columns")
    print("---- List columns (displayName -> name) ----")
    for c in cols.get("value", []):
        print(f"{c.get('displayName')} -> {c.get('name')}")


# -------------------- Main --------------------
def main():
    graph_token = get_graph_token()

    # Email mailbox that receives the report
    mailbox_upn = "apatil@glassdoctordfw.com"

    # SharePoint list location
    hostname = "glassdoctornt.sharepoint.com"
    site_path = "/sites/Retail"
    list_name = "Retail Problem/Incomplete Ops Job"

    site_id, list_id = get_site_and_list_ids(graph_token, hostname, site_path, list_name)
    #print_list_columns(graph_token, site_id, list_id)
    #return

    # Find latest email
    msg = latest_message_for_subject(graph_token, mailbox_upn, SUBJECT)
    if not msg:
        print(f"No email found with subject containing: {SUBJECT}")
        return

    # Get xlsx attachment
    fname, xlsx = get_first_xlsx_attachment_from_message(graph_token, mailbox_upn, msg["id"])
    if not xlsx:
        print("Email found but no .xlsx attachment.")
        return

    # Read rows
    rows = read_xlsx_first_sheet_rows(xlsx)
    header = rows[0]
    body = rows[1:]

    # Detect columns in attachment
    bu_col   = find_col_idx(header, ["business unit"])
    job_col  = find_col_idx(header, ["job #", "job number", "job"])
    cust_col = find_col_idx(header, ["customer name"])
    sub_col  = find_col_idx(header, ["jobs subtotal", "subtotal", "job amount"])
    date_col = find_col_idx(header, ["next appt start date", "next appt date"])

    if None in (bu_col, job_col, cust_col, sub_col, date_col):
        raise RuntimeError(f"Missing required columns. Header detected: {header}")

    # Build map of existing list items by Title (Job Number)
    job_field = FIELDS["job_number"]  # "Title"
    existing_map = load_existing_job_map(graph_token, site_id, list_id, job_field)
    print(f"Loaded {len(existing_map)} existing items into job map.")

    upserts = 0
    for r in body:
        if len(r) <= max(bu_col, job_col, cust_col, sub_col, date_col):
            continue

        bu_raw = normalize_key(r[bu_col])
        job_number = str(r[job_col]).strip()
        if not job_number:
            continue

        assigned_email = ASSIGNED_MAP.get(bu_raw)
        if not assigned_email:
            continue

        next_dt = try_parse_excel_date(r[date_col])

        fields_payload = {
            FIELDS["job_number"]: job_number,
            FIELDS["customer_name"]: str(r[cust_col]).strip(),
            FIELDS["job_amount"]: parse_money(r[sub_col]),
            FIELDS["assigned_to"]: assigned_email,  # TEXT column
        }
        if next_dt:
            fields_payload[FIELDS["next_appt_date"]] = next_dt

        existing_id = existing_map.get(job_number)

        if not existing_id:
            # Create minimal item first (Title only) to avoid Graph flakiness
            created = create_item(graph_token, site_id, list_id, {FIELDS["job_number"]: job_number})
            new_id = created["id"]

            # Patch remaining fields (excluding Title)
            patch_fields = dict(fields_payload)
            patch_fields.pop(FIELDS["job_number"], None)

            if patch_fields:
                update_item_fields(graph_token, site_id, list_id, new_id, patch_fields)

            existing_map[job_number] = new_id
        else:
            update_item_fields(graph_token, site_id, list_id, existing_id, fields_payload)

        upserts += 1

    print(f"Done. Processed attachment: {fname}. Upserted rows: {upserts}")

if __name__ == "__main__":
    main()
